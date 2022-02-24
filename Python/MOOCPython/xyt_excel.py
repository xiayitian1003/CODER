import openpyxl
import XmlWithExcel
from xythttprequest import pick_point_search_for_city, pick_point_reverse_geo_code
from xythttprequest import gaode_reverse_geo_code
from xythttprequest import gaode_search_for_city
from decimal import Decimal
import time


# Path = "D:\\CODE\\workstation\\Document\\NewestTimeZone.xlsx"
# Path = "D:\\CODE\\workstation\\Document\\9.0\\TimeZone-Version-2.0.xlsx"
# Path = "D:\\CODE\\workstation\\Document\\9.0\\v2.0NotFound.xlsx"
# Path = "D:\\CODE\\workstation\\Document\\9.0\\TimeZone-Version-2.0-Reverse.xlsx"


# print(write_sheet.title)
# print(write_sheet.min_row, write_sheet.max_row)
# print(write_sheet.min_column, write_sheet.max_column)


def round_3_float(decimal_num):
    decimal_num = str(decimal_num)
    return Decimal(decimal_num).quantize(Decimal('0.000'))


def get_cell_value(cell):
    return cell.value


# def search_all_city():
#     eng_excel_dir = "D:\\CODE\\workstation\\Document\\9.0\\convert2Eng2Translate.xlsx"
#     book = openpyxl.load_workbook(eng_excel_dir)
#     sheet = book.worksheets[1]
#     write_sheet = book.create_sheet()
#     print(sheet.title)
#     print(sheet.min_row, sheet.max_row)
#     print(sheet.min_column, sheet.max_column)
#     unfinished_city_sheet = book.create_sheet()
#     cnt = int(0)
#     for row in sheet.rows:
#         cnt += 1
#         print("Start search %s in %d times" % (row[0].value, cnt))
#         city_info = (row[0].value, row[1].value)
#         ans = pick_point_search_for_city(row[0].value, row[1].value)
#         ans_row = city_info
#         if ans is not None:
#             print("%d times finish, ans of %s is %.3f, %.3f" % (cnt, row[0].value, ans['lat'], ans['lon']))
#             ans_row += (round_3_float(ans['lat']), round_3_float(ans['lon']), "", ans['display_name'])
#             write_sheet.append(ans_row)
#         else:
#             print("%d times finish, ans of %s is None" % (cnt, row[0].value))
#             ans_none_type = "city&country"
#             ans = pick_point_search_for_city(row[0].value, None)
#             unfinished_ans = (cnt,) + city_info
#             if ans is not None:
#                 print("%3d %10s %10s %10s" % (cnt, row[0].value, row[1].value, row[2].value))
#                 ans_row += (ans['lat'], ans['lon'], ans_none_type, ans['display_name'])
#                 unfinished_ans += (ans['lat'], ans['lon'], ans_none_type, ans['display_name'])
#             else:
#                 print("city")
#                 ans_lat, ans_lon = "none", "none"
#                 ans_none_type = "city"
#                 ans_row += (ans_lat, ans_lon, ans_none_type, "none name")
#                 unfinished_ans += (ans_lat, ans_lon, ans_none_type, "none name")
#
#             write_sheet.append(ans_row)
#             unfinished_city_sheet.append(unfinished_ans)
#         book.save(filename=eng_excel_dir)
#
#
def search_by_reverse():
    eng_excel_dir = "D:\\CODE\\workstation\\Document\\9.0\\convert2Eng2Translate.xlsx"
    book = openpyxl.load_workbook(eng_excel_dir)
    sheet = book.worksheets[2]
    write_sheet = book.create_sheet()
    print(sheet.title)
    print(sheet.min_row, sheet.max_row)
    print(sheet.min_column, sheet.max_column)
    cnt = int(0)
    for row in sheet.rows:
        cnt += 1
        print("Start search %s in %d times" % (row[0].value, cnt))
        print("coordinate is -- lat: %s, lon: %s" % (row[2].value, row[3].value))

        # print(row, row[0].col_idx, row[0].column, row[0].column_letter, row[0].coordinate, row[0].row)
        ans = pick_point_reverse_geo_code(row[2].value, row[3].value)
        print(ans)

        ans_row = ()
        ans_row = ans_row + tuple(map(get_cell_value, row[0: 4]))
        print(ans_row)
        plan_to_add = [ans['city'], ans['country'], ans['country_code']]
        print(type([ans['city'], ans['country'], ans['country_code']]))
        for item in plan_to_add:
            if item is not None and len(item) > 0:
                ans_row += (item,)
        ans_row += (row[5].value,)

        write_sheet.append(ans_row)
        book.save(filename=eng_excel_dir)


def get_final_wrong_list():
    final_wrong_path = "D:\\CODE\\workstation\\Document\\9.0\\FinalWrongTimeZone.xlsx"
    book = openpyxl.load_workbook(final_wrong_path)

    sheet = book.worksheets[2]
    final_asure = book.create_sheet()

    cnt = int(0)
    num = cnt
    for row in sheet.rows:
        cnt += 1
        print(cnt)
        if str(row[5].value) == "1":
            num += 1
            ans_row = tuple(map(get_cell_value, row))
            ans = gaode_reverse_geo_code(row[8].value, row[7].value)
            plan_to_add = (
                str(ans['city']), str(ans['country']), str(ans['countrycode']), str(ans['formatted_address']))
            ans_row += plan_to_add
            final_asure.append(ans_row)
        time.sleep(0.1)
    print(num)
    book.save(filename=final_wrong_path)


def correct_location(city, lat, lon, sheet):
    cnt, num = int(0), int(0)
    for row in sheet.rows:
        cnt += 1
        lat_xy, lon_xy = 'H' + str(cnt), 'I' + str(cnt)
        if str(row[5].value) == "1":
            if city == row[1].value:
                sheet[lat_xy], sheet[lon_xy] = lat, lon
                print(city, lat, lon, cnt, num)


def back_to_final_wrong():
    final_wrong_path = "D:\\CODE\\workstation\\Document\\9.0\\FinalWrongTimeZone.xlsx"
    need_to_translate_dir = "D:\\CODE\\workstation\\Document\\9.0\\convert2Eng2Translate.xlsx"
    book, translation_book = openpyxl.load_workbook(final_wrong_path), openpyxl.load_workbook(need_to_translate_dir)

    sheet, trans_sheet = book.worksheets[0], translation_book.worksheets[0]
    # sheet['K1'] = "K1!"
    cnt = int(0)
    for row in trans_sheet.rows:
        cnt += 1
        print(cnt, row[0].value, row[2].value, row[3].value)
        correct_location(row[0].value, row[2].value, row[3].value, sheet)

    print(cnt)
    book.save(filename=final_wrong_path)


def get_excel_city_set():
    raw_city_list_dir = "D:\\CODE\\workstation\\Document\\9.0\\TimeZone-Version-2.0.xlsx"
    # raw_city_list_dir = "D:\CODE\workstation\Document\9.0\TimeZone-Version-2.0.xlsx"
    book = openpyxl.load_workbook(raw_city_list_dir)
    raw_city_sheet = book.worksheets[0]
    excel_city_set = set()
    for row in raw_city_sheet.rows:
        excel_city_set.add(row[0].value)
    print("excel size is :", len(excel_city_set))
    return excel_city_set


def get_translate_city_list():
    need_to_translate_dir = "D:\\CODE\\workstation\\Document\\9.0\\convert2Eng2Translate.xlsx"
    book = openpyxl.load_workbook(need_to_translate_dir)
    translate_city_sheet = book.worksheets[0]
    translate_list = []
    for row in translate_city_sheet.rows:
        if row[0].value is not None and len(row[0].value) > 0:
            translate_list.append(row[0].value)
    print(len(translate_list))
    return translate_list


def write_translation():
    ans_lst = XmlWithExcel.translated_city_list()
    translated_ans = "D:\\CODE\\workstation\\Document\\9.0\\convert2Eng2Translate.xlsx"
    book = openpyxl.load_workbook(translated_ans)
    ans_sheet = book.create_sheet()
    for trans in ans_lst:
        ans_sheet.append((trans,))
    book.save(filename=translated_ans)


def main():
    # pick_point_search_for_city("Reno", "united states")
    back_to_final_wrong()
    # pick_point_reverse_geo_code(28.536, -15.749)
    # search_all_city()
    # search_by_reverse()


# print("excel main thread is running!!!")
# get_excel_city_set()

# get_final_wrong_list()
# search_by_reverse()
# search_all_city()


if __name__ == '__main__':
    main()
